import sys
import re
import os
import math
import pandas as pd
import numpy as np
import plotly.express as px
import datetime as dt

from unidecode import unidecode
import nltk
from xlrd import open_workbook
import openpyxl
from fuzzywuzzy import fuzz
from fuzzywuzzy import process

pd.set_option('max_colwidth', None)
pd.set_option('display.max_rows', 500)

def gendf_from_excel_table(input_file, table_cols, stop_if_empty=True):
    name, extension = os.path.splitext(input_file)
    if extension == ".xlsx":
        wb_ = openpyxl.load_workbook(input_file)
        wb = wb_.active
        row_range = range(1, wb.max_row+1)
        col_range = range(1,wb.max_column+1)
    elif extension == ".xls":
        wb = open_workbook(input_file)
        sheet = wb.sheets()[0] #frist sheet
        row_range = range(sheet.nrows)
        col_range = range(sheet.ncols)
    else:
        print("Error, unsoported excel extension when reading file: " + input_file)
        return
    
    table_rows = []
    valid_row = False
    for row in row_range:
        rowl = []
        for col in col_range:
            if extension == ".xlsx":
                value = wb.cell(row=row,column=col).value
            elif extension == ".xls":
                #print("{}:{}".format(row,col))
                value = sheet.cell(row,col).value
            if value == None:
                value = np.nan
            try:
                rowl.append(unidecode(value))
            except:
                rowl.append(value)

            if all(item in rowl for item in table_cols): #we found the headers
                valid_row = True
             
        if stop_if_empty and valid_row and (rowl.count(np.nan) + rowl.count(''))/len(rowl) > 0.9: #stop if more than 90% of cells are empty
            break

        if valid_row:
            table_rows.append(rowl)
    
    #create dataframe
    df = pd.DataFrame(table_rows).dropna(axis=0, how="all").dropna(axis=1, how="all")
    if not df.empty:
        df.columns = df.iloc[0]
        df = df.tail(-1)
    return df
    