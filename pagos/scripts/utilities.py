import sys
import re
import os
import math
import pandas as pd
import numpy as np
import plotly.express as px
import datetime as dt

from unidecode import unidecode
import nltk
from xlrd import open_workbook
import xlrd
import openpyxl
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
from openpyxl.cell.cell import Cell

pd.set_option('max_colwidth', None)
pd.set_option('display.max_rows', 500)

def is_interactive():
    import __main__ as main
    return not hasattr(main, '__file__')

def gendf_from_excel_table(input_file, table_cols, stop_if_empty=True):
    name, extension = os.path.splitext(input_file)
    if extension == ".xlsx":
        wb_ = openpyxl.load_workbook(input_file)
        wb = wb_.active
        row_range = range(1, wb.max_row+1)
        col_range = range(1,wb.max_column+1)
    elif extension == ".xls":
        wb = open_workbook(input_file)
        sheet = wb.sheets()[0] #frist sheet
        row_range = range(sheet.nrows)
        col_range = range(sheet.ncols)
    else:
        print("Error, unsoported excel extension when reading file: " + input_file)
        return
    
    table_rows = []
    valid_row = False
    prev_rowl_count = 0
    for row in row_range:
        rowl = []
        for col in col_range:
            if extension == ".xlsx":
                cell = wb.cell(row=row, column=col)
                if isinstance(cell, Cell):
                    if cell.is_date:
                        value = cell.value.strftime("%d/%m/%Y")
                    else:
                        value = cell.value
                else:
                    value = ""
            elif extension == ".xls":
                cell = sheet.cell(row, col)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    date_value = xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode)
                    value = date_value.strftime("%d/%m/%Y")
                else:
                    value = cell.value


            if value == None:
                value = np.nan
            try:
                rowl.append(unidecode(value))
            except:
                rowl.append(value)

            if all(item in rowl for item in table_cols): #we found the headers
                valid_row = True
             
        #if stop_if_empty and valid_row and (rowl.count(np.nan) + rowl.count(''))/len(rowl) > 0.9: #stop if more than 90% of cells are empty
        if stop_if_empty and valid_row and (len(rowl) - rowl.count(np.nan) - rowl.count('')) < prev_rowl_count -1:
            if row < 0.8 * row_range[-1]:
                print(f"Error: se procesaron muy pocas lineas. Revise si la tabla del archivo '{input_file}' tiene errors como columnas desalineadas o celdas extra.")
                exit(-1)
            break
        if len(table_rows) > 0:
            prev_rowl_count = len(rowl) - rowl.count(np.nan) - rowl.count('')
        
        if valid_row:
            table_rows.append(rowl)

    #create dataframe
    df = pd.DataFrame(table_rows).dropna(axis=0, how="all").dropna(axis=1, how="all")
    if not df.empty:
        df.columns = df.iloc[0]
        #df = df.tail(-1)
    return df
    
